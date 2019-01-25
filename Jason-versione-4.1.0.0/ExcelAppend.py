from os.path import join
import openpyxl 
import logging


logging.basicConfig(format = '%(asctime)s %(message)s',
                            datefmt = '%m/%d/%Y %I:%M:%S %p',
                            filename = 'system.log',
                            level=logging.INFO)


def create_excel(dict_path, files_list, name_file_output):
    try:
        excel_template = openpyxl.load_workbook(join(dict_path["excel_templates_path"], "template_leonardo.xlsx"))
        sheet_template = excel_template.active
        number_cell = 7
        for file_checkbox in files_list:
            if(file_checkbox.isChecked() == True):
                excel_document = openpyxl.load_workbook(join(dict_path["excel_files_path"], file_checkbox.text()))
                sheet = excel_document.active
                row = 7
                cols = {1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G'}
                while(sheet.cell(row=row, column=1).value != None):  
                    for col_number, col_letter in cols.items():
                        sheet_template[col_letter+str(number_cell)] = sheet.cell(row=row, column=col_number).value
                    number_cell+=1 
                    row += 1
                row = 7   
        excel_template.save(join(dict_path["excel_final_path"], name_file_output+".xlsx"))
        return True
    except:
        return False